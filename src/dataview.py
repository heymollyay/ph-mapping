import os
import customtkinter
import tkinter as tk
import openpyxl
import csv
from tkinter import filedialog, ttk
from src.config import NEW_PROJECT_TITLE, STICKY_NSEW, DEFAULT_PADX, DEFAULT_PADY

class DataView(customtkinter.CTkFrame):
    def __init__(self, master):
        super().__init__(master)

        self.tree = ttk.Treeview(self)
        self.tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.v_scroll = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        self.h_scroll = ttk.Scrollbar(self.tree, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)
        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        self.open_file = customtkinter.CTkButton(self, text="Open File", command=self.get_file)
        self.open_file.pack()

    def get_file(self):
        filetypes = [
            ("Supported file types", "*.xlsx;*.xlsm;*.csv"),
            ("Excel files", "*.xlsx;*.xlsm"),
            ("CSV files", "*.csv"),
            ("All files", "*.*")
        ]

        self.my_file = filedialog.askopenfilename(initialdir="/", title= "Select File", filetypes=filetypes)

        if self.my_file:
            self.display_file()

    def display_file(self):
        if not self.my_file:
            return

        self.tree.delete(*self.tree.get_children())
        self.tree["columns"]= ()

        file_ext = os.path.splitext(self.my_file)[1].lower()
        try:
            if file_ext in (".xlsx", ".xlsm"):
                self.display_excel_openpyxl()
            elif file_ext == ".csv":
                self.display_csv()
            else:
                print("Unsupported file format")
        except Exception as e:
            print(f"Error loading file: {str(e)}")

    def display_excel_openpyxl(self):
        workbook = openpyxl.load_workbook(self.my_file)
        worksheet = workbook.active

        # Get headers (first row)
        headers = [cell.value if cell.value else f"Col{idx + 1}" for idx, cell in
                   enumerate(next(worksheet.iter_rows(min_row=1, max_row=1)))]
        self.tree["columns"] = headers
        for col in headers:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)

        # Load data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
            values = [str(cell) if cell is not None else "" for cell in row]
            self.tree.insert("", "end", text=f"Row {row_idx}", values=values)

    def display_csv(self):
        with open(self.my_file, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)  # First row as headers
            self.tree["columns"] = headers
            for col in headers:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100)

            for row_idx, row in enumerate(reader, start=1):
                self.tree.insert("", "end", text=f"Row {row_idx}", values=row)



